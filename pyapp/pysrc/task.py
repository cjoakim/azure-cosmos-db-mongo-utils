import json
import os
import os.path
import shutil
import time
import traceback
import warnings

import arrow
from bson import json_util
from openpyxl import load_workbook

from pysrc.artifacts import ArtifactGenerator
from pysrc.bytes import Bytes
from pysrc.aggregators import ContainersAggregator
from pysrc.datasets import Datasets
from pysrc.docscan import DocscanCluster
from pysrc.docscan import DocscanClusterResult
from pysrc.env import Env
from pysrc.mma_files import MmaFiles
from pysrc.fs import FS
from pysrc.indices import Indices
from pysrc.mongo import Mongo
from pysrc.reporter import Reporter
from pysrc.shards import Shards
from pysrc.system import System

# Class Tasks defines common application logic that can be executed either
# from the console app (main.py) or the web app (web.py).
# Class TaskResult is also defined at the end of this file.
#
# Chris Joakim, Microsoft, 2023

class Tasks(object):

    def __init__(self):
        pass

    @classmethod
    def check_env(cls):
        tr = TaskResult('check_env')
        try:
            Datasets.reset()
            home = Env.var('HOMEPATH')
            tr.add_message('HOMEPATH: {}'.format(home))
            tr.add_message('USERNAME: {}'.format(Env.username()))
            tr.add_message('pwd:               {}'.format(FS.pwd()))
            tr.add_message('mma_output_dir:    {}'.format(Datasets.mma_output_dir()))
            tr.add_message('mma_logs_dir:      {}'.format(Datasets.mma_logs_dir()))
            tr.add_message('customer_data_dir: {}'.format(Datasets.customer_data_dir()))
            tr.set_success(home is not None)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def despace_file(cls, infile, outfile):
        tr = TaskResult('despace_file')
        # Typical use: Azure Data Studio creates query result csv export files
        # with unnecessary space lines.  These can be removed here.
        try:
            Datasets.reset()
            tr.add_message('infile: {}'.format(infile))
            tr.add_message('outfile: {}'.format(outfile))
            lines, out_lines = FS.read_lines(infile), list()
            for line in lines:
                stripped = line.strip()
                if len(stripped) > 0:
                    out_lines.append(stripped)
            FS.write_lines(out_lines, outfile)
            tr.add_message('input line count: {}'.format(len(lines)))
            tr.add_message('output line count: {}'.format(len(out_lines)))
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def trello_reports(cls, infile):
        # https://developer.atlassian.com/cloud/trello/rest/api-group-actions/

        tr = TaskResult('trello_reports')
        try:
            tr.log('trello_reports, infile: {}'.format(infile))
            t = Trello(infile)
            t.produce_reports()


            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def migration_wave_report(cls):
        tr = TaskResult('migration_wave_report')
        try:
            r = Reporter()
            r.migration_wave_report()
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def gen_mma_execution_scripts(cls, clusters):
        tr = TaskResult('gen_mma_execution_scripts')
        try:
            unique_wave_numbers = dict()
            for cluster_name, cluster_url in clusters.items():
                wave = int(cluster_name.split('-')[0])
                unique_wave_numbers[wave] = ''
            FS.write_json(unique_wave_numbers, 'tmp/unique_wave_numbers.json')

            # Produce an MMA execution script for EACH wave
            all_cluster_names = list()
            for wave_num in sorted(unique_wave_numbers):
                wave_cluster_names = cls.cluster_names_for_wave(clusters, wave_num)
                for name in wave_cluster_names:
                    all_cluster_names.append(name)
                cls.gen_mma_execution_script(clusters, wave_num, wave_cluster_names)

            # Produce an MMA execution script for ALL waves (i.e. - wave 1000), in wave sequence
            cls.gen_mma_execution_script(clusters, 1000, all_cluster_names)
            tr.log("Note: copy the appropriate mma_nnn.ps1 file to the directory where MongoMigrationAssessment.exe")
            tr.log("is located on your system.")
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def cluster_names_for_wave(cls, clusters, wave_num):
        names_for_wave = list()
        for cluster_name in sorted(clusters.keys()):
            wave = int(cluster_name.split('-')[0])
            if wave == wave_num:
                names_for_wave.append(cluster_name)
        return sorted(names_for_wave)

    @classmethod
    def gen_mma_execution_script(cls, clusters, wave_num, wave_cluster_names):
        script_lines = list()
        redirect_dir = '{}/current/mmaout'.format(FS.pwd())
        #print('redirect_dir: {}'.format(redirect_dir))

        # collect the unique urls
        unique_urls, visited_urls = dict(), dict()
        for cluster_name in wave_cluster_names:
            url = clusters[cluster_name]['source']
            if url in unique_urls.keys():
                unique_urls[url].append(cluster_name)
            else:
                unique_urls[url] = [cluster_name]

        script_lines.append('')
        script_lines.append('# Wave Number:      {}'.format(wave_num))
        script_lines.append('# Cluster Count:    {}'.format(len(wave_cluster_names)))
        script_lines.append('# Unique URL Count: {}'.format(len(unique_urls)))

        for cluster_idx, cluster_name in enumerate(sorted(wave_cluster_names)):
            conn_str = clusters[cluster_name]['source']
            if conn_str in visited_urls.keys():
                script_lines.append('')
                script_lines.append('# cluster {} has a redundant duplicate URL, see clusters: {}'.format(
                    cluster_name, unique_urls[conn_str]))
            else:
                visited_urls[conn_str] = ''
                script_lines.append('')
                script_lines.append('date')
                script_lines.append("Write-Output 'Executing MMA for {} -> {}'".format(cluster_idx, cluster_name))
                line = '.\MongoMigrationAssessment.exe -c "{}" -t 4.2 -n "{}" | Out-File -Encoding utf8 -FilePath {}/{}-{}-mma.txt'.format(
                    conn_str, cluster_name, redirect_dir, cluster_name, cluster_idx)
                script_lines.append(line)

        script_lines.append('')
        script_lines.append('date')
        script_lines.append("Write-Output 'Done'")
        user = Env.username()
        FS.write_lines(script_lines, 'tmp/mma_{}.ps1'.format(wave_num))

    @classmethod
    def gen_spark_migration_config_files(cls):
        tr = TaskResult('gen_spark_migration_config_files')
        try:
            tr.log('gen_spark_migration_config_files not yet implemented')
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def gen_cosmos_provisioning_scripts(cls):
        tr = TaskResult('gen_cosmos_provisioning_scripts')
        try:
            agen = ArtifactGenerator()
            agen.generate_bicep_and_powershell_artifacts()
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def aggregate_mma_execution_output(cls):
        tr = TaskResult('aggregate_mma_execution_output')
        try:
            mma_output_dir = Datasets.mma_output_dir()
            dirs = FS.list_directories_in_dir(mma_output_dir)
            dirs_dict = dict()
            for dir in dirs:
                dirs_dict[dir] = ''

            # walk the MMA output directory and collect all file info and filenames
            # then validate the presence/absence of key expected files
            tr.log('mma_output_dir: {}'.format(mma_output_dir))
            mma_output_files = FS.walk(mma_output_dir)
            Datasets.write_mma_output_dir_walk_list(mma_output_files)
            Datasets.write_mma_output_dir_walk_filenames(mma_output_files)
            cls.validate_presence_of_expected_mma_files(dirs_dict, mma_output_files)

            aggregated_json_files = list()
            tr.log("{} mma output files".format(len(mma_output_files)))
            cluster_mapping = dict()

            # First read/parse mongo_migration_assessment_report.html to associate cluster with random uuid.
            for file_idx, file_info in enumerate(mma_output_files):
                if file_info['base'] == 'mongo_migration_assessment_report.html':
                    name_id_tup = cls.__parse_assessment_name_and_id(file_info)
                    name, id = name_id_tup[0], name_id_tup[1]
                    cluster_mapping[id] = '{}'.format(name)

            if len(cluster_mapping) < 1:
                # no mongo_migration_assessment_report.html files were found; therefore MMA was generated by ADS
                # use the top-level uuid directory name as the cluster name
                for file_idx, file_info in enumerate(mma_output_files):
                    try:
                        if file_info['base'] == 'instance_detail.json':
                            tokens = file_info['dir'].replace("/"," ").replace("\\", " ").strip().split(' ')
                            for token_idx, token in enumerate(tokens):
                                if token == 'MongoMigrationAssessment':
                                    uuid_str = tokens[token_idx + 1]
                                    cluster_mapping[uuid_str] = uuid_str
                                    print('cluster_mapping added based on instance_detail.json: {}'.format(uuid_str))
                    except:
                        print('exception parsing for instance_detail.json instead of mongo_migration_assessment_report.html')
            else:
                print('{} cluster names parsed from the mongo_migration_assessment_report.html file(s)'.format(len(cluster_mapping)))

            # Next, iterate the JSON files and collect each
            for file_idx, file_info in enumerate(mma_output_files):
                fqname = file_info['full']
                if fqname.endswith('.json'):
                    obj = dict()
                    obj['_file_idx'] = file_idx
                    obj['_file_name'] = fqname
                    obj['_file_basename'] = file_info['base']
                    epoch = os.path.getmtime(fqname)
                    obj['_file_mtime'] = epoch
                    obj['_file_date'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(epoch))
                    if Env.verbose():
                        tr.log('reading json file: {}'.format(fqname))
                    obj['_cluster'] = cls.__lookup_cluster_name(cluster_mapping, file_info['abspath'])
                    obj['data'] = FS.read_json(fqname)
                    aggregated_json_files.append(obj)

            verification_dict = dict()
            for dir in dirs_dict.keys():
                if dir in cluster_mapping.keys():
                    verification_dict[dir] = 'present'
                else:
                    verification_dict[dir] = 'absent'
                    cls.__verify_mongo_migration_assessment_report_html_file(dir, mma_output_files)

            Datasets.write_aggregated_mma_output_file(aggregated_json_files)
            Datasets.write_cluster_uuid_mappings_file(cluster_mapping)
            Datasets.write_cluster_mappings_verification_file(verification_dict)
            Datasets.display()
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def validate_presence_of_expected_mma_files(cls, dirs_dict, mma_output_files):
        for parent_dir in sorted(dirs_dict.keys()):
            mmaf = MmaFiles(parent_dir)
            for file_obj in mma_output_files:
                if parent_dir in file_obj['abspath']:
                    mmaf.add(file_obj)

            print(json.dumps(mmaf.get_data(), sort_keys=False, indent=2))
            if mmaf.all_present():
                print("expected_files_present in assessment dir: {}".format(parent_dir))
            else:
                print("expected_files_absent in assessment dir:  {}".format(parent_dir))

    @classmethod
    def capture_mma_logs(cls):
        tr = TaskResult('capture_mma_logs')
        try:
            logs_dir = Datasets.mma_logs_dir()
            tr.log('logs_dir: {}'.format(logs_dir))
            files = FS.list_files_in_dir(logs_dir)
            for file in files:
                source = '{}/{}'.format(logs_dir, file)
                target = 'tmp/{}'.format(file)
                tr.log('copying {} --> {}'.format(source, target))
                shutil.copyfile(source, target)
                tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def gather_cluster_db_collection_info(cls):
        tr = TaskResult('gather_cluster_db_collection_info')
        try:
            aggregated_mma_outputs = Datasets.read_aggregated_mma_output_file()
            agg = ContainersAggregator(aggregated_mma_outputs)
            containers_data_list = agg.containers_obj.get_data_list()
            Datasets.write_aggregated_containers_info_file(containers_data_list)
            tr.log('containers count: {}'.format(len(containers_data_list)))

            overrides_data = agg.collect_overrides_data(containers_data_list)
            Datasets.write_mappings_generated_config_file(overrides_data)
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def gather_shard_info(cls):
        tr = TaskResult('gather_shard_info')
        try:
            aggregated_mma_outputs = Datasets.read_aggregated_mma_output_file()
            Shards(aggregated_mma_outputs).extract_keys_and_advice()
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def gather_index_info(cls):
        tr = TaskResult('gather_index_info')
        try:
            aggregated_mma_outputs = Datasets.read_aggregated_mma_output_file()
            Indices(aggregated_mma_outputs).extract_info_and_advice()
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def pymongo_connect(cls, cluster_name, conn_str):
        tr = TaskResult('pymongo_connect')
        try:
            lines = list()
            host = conn_str.split('@')[1]
            lines.append('pymongo_connect, cluster: {} host: {}'.format(cluster_name, host))
            outfile = 'mmaout/{}-pymongo.txt'.format(cluster_name)
            opts = dict()
            opts['conn_string'] = conn_str
            opts['verbose'] = False
            m = Mongo(opts)
            bypass_dbnames = ['admin', 'config', 'local']
            dbs = m.list_databases()
            lines.append('cluster_name: {} outfile: {}'.format(cluster_name, outfile))
            if dbs is not None:
                lines.append('cluster:    {} has {} databases: {}'.format(cluster_name, len(dbs), dbs))
                for dbname in dbs:
                    if dbname in bypass_dbnames:
                        pass
                    else:
                        m.set_db(dbname)
                        colls = m.list_collections()
                        if colls is not None:
                            lines.append('database:   {} has {} collections: {}'.format(dbname, len(colls), colls))
                            for coll in colls:
                                lines.append(
                                    'collection: {}|{}|{}|{}'.format(cluster_name, host, dbname, coll))  # grep this
                        else:
                            lines.append('database:   {} has {} no collections'.format(dbname))
            else:
                lines.append('cluster:    {} has no databases'.format(cluster_name))
            FS.write_lines(lines, outfile)
            for line in lines:
                tr.add_message(line)
            tr.log('file written: {}'.format(outfile))
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def read_parse_clusters_info_excel_file(cls):
        tr = TaskResult('read_parse_clusters_info_excel_file')
        try:
            infile = 'current/customer-source-dest-spreadsheet.xlsx'
            wsname = 'Sheet1'
            clusters = dict()
            raw_urls = list()
            tr.set_resp_obj(clusters)
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            wb = load_workbook(filename=infile)
            print('sheetnames: '.format(wb.sheetnames))
            tr.log(wb.sheetnames)
            ws = wb[wsname]
            col_count = ws.max_column
            row_count = ws.max_row
            mdb_cred  = FS.read_single_line('current/cred/mdb-cred.txt')
            print('mdb_cred: {}'.format(mdb_cred))
            tr.log('worksheet: {}, cols: {}, rows: {}'.format(wsname, col_count, row_count))
            for r in range(2, row_count):
                # ugh, 'org' is now a calculated field in the spreadsheet
                prty   = str(ws.cell(row=r, column=3).value).replace(' ', '').strip()  # column C
                bar    = str(ws.cell(row=r, column=4).value).replace(' ', '').strip()  # column D
                env    = str(ws.cell(row=r, column=5).value).replace(' ', '').strip()  # column E
                cname  = str(ws.cell(row=r, column=6).value).replace(' ', '').strip()  # column F
                cosmos = str(ws.cell(row=r, column=28).value).replace(' ', '').strip()  # column AB
                vcore  = str(ws.cell(row=r, column=29).value).replace(' ', '').strip()  # column AC
                #org   = '{}-{}-{}---{}'.format(prty, bar, env, cname).strip()
                org    = '1-{}-{}---{}'.format(bar, env, cname).strip()  # use 1 for all waves
                url    = str(ws.cell(row=r, column=12).value).replace(' ', '').strip()
                raw_urls.append(url)

                mapping = dict()
                mapping['source'] = ''
                mapping['target'] = cosmos

                if cls.org_is_valid(org):
                    print('ORD VALID - org: {}  url: {}'.format(org, url))
                    if cls.url_is_valid(url):
                        if mdb_cred in url:
                            # the url from the spreadsheet already contains the credentials
                            mapping['source'] = url
                            clusters[org] = mapping
                        else:
                            # inject the read-only-credentials into the URL to form the connection string
                            url_tokens = url.split('//')
                            conn_str = '{}//{}@{}'.format(url_tokens[0], mdb_cred, url_tokens[1])
                            mapping['source'] = conn_str
                            clusters[org] = mapping
                    else:
                        print('URL INVALID: {}'.format(url))
                else:
                    print('ORD INVALID: {}'.format(org))

            outfile = 'current/excel_clusters.json'
            FS.write_json(clusters, outfile)
            tr.log('file written: {} with {} clusters'.format(outfile, len(clusters)))
            FS.write_json(raw_urls, 'tmp/excel_raw_urls.json')

            # check for duplicate urls
            unique_urls, duplicate_count, total_count = dict(), 0, len(clusters)
            for key, value in clusters.items():
                source = value['source']
                if source in unique_urls.keys():
                    unique_urls[source].append(key)
                    duplicate_count = duplicate_count + 1
                else:
                    unique_urls[source] = [key]
                print('CLUSTER {} -> {}'.format(key, value))
            print('CLUSTER_TOTALS - total: {}, duplicate_urls: {}'.format(total_count, duplicate_count))
            FS.write_json(unique_urls, 'current/excel_unique_cluster_urls.json')
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def org_is_valid(cls, s):
        if s is None:
            return False
        if 'None' in s:
            return False
        if len(s) < 1:
            return False
        if 'CLUSTERNAME' in s:
            return False
        return True

    @classmethod
    def url_is_valid(cls, s):
        if s is None:
            return False
        if 'None' in s:
            return False
        if len(s) < 1:
            return False
        if 'ServerAddress' in s:
            return False
        if s.startswith('mongodb+srv://'):
            return True
        else:
            return False

    @classmethod
    def gen_pymongo_connect_script(cls, clusters):
        tr = TaskResult('gen_pymongo_connect_script')
        try:
            script_lines = list()
            script_lines.append('')
            script_lines.append('# Execute a python/pymongo client vs known clusters from Excel file.')
            script_lines.append('# Get the list of databases and their collections in each cluster.')
            script_lines.append('# UTC Date Generated: {}'.format(arrow.utcnow().format('YYYY-MM-DD HH:mm:ss ZZ')))
            script_lines.append('# Chris Joakim, Microsoft, swift-utils')

            for cluster_idx, cluster_name in enumerate(sorted(clusters.keys())):
                conn_str = clusters[cluster_name]['source']
                script_lines.append('')
                script_lines.append(
                    "Write-Output 'Executing pymongo connect for {} -> {}'".format(cluster_idx, cluster_name))
                line = 'python main.py pymongo_connect {} "{}"'.format(cluster_name, conn_str)
                script_lines.append(line)
            script_lines.append('')
            script_lines.append("Write-Output 'Done'")
            outfile = 'pymongo_connect.ps1'
            FS.write_lines(script_lines, outfile)
            tr.log('file written: {}'.format(outfile))
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def scan_mmaout_results(cls):
        tr = TaskResult('scan_mmaout_results')
        try:
            results = dict()
            tr.set_resp_obj(results)
            clusters = FS.read_json('current/excel_clusters.json')
            filenames = list()
            for file_obj in FS.walk('current/mmaout'):
                filenames.append('current/mmaout/{}'.format(file_obj['base']))

            for cluster_name in sorted(clusters.keys()):
                mma_file = cls._find_file_in_list(filenames, cluster_name, 'mma.txt')
                pymongo_file = cls._find_file_in_list(filenames, cluster_name, 'pymongo.txt')
                mma_file_key = '{} mma_present'.format(cluster_name)
                mma_successful_key = '{} mma_successful'.format(cluster_name)
                pymongo_file_key = '{} pymongo_present'.format(cluster_name)
                database_count, collection_count = 0, 0
                results[cluster_name] = 'cluster'
                results[mma_file_key] = mma_file
                results[mma_successful_key] = 'unknown'
                results[pymongo_file_key] = str(pymongo_file)

                if pymongo_file is not None:
                    lines = FS.read_lines(pymongo_file)
                    pymongo_databases_key = '{} pymongo_databases'.format(cluster_name)
                    pymongo_collections_key = '{} pymongo_collections'.format(cluster_name)
                    for line in lines:
                        scrubbed = line.strip()
                        if scrubbed.find('database:') >= 0:
                            database_count = database_count + 1
                        if 'collection:' in line:
                            collection_count = collection_count + 1
                    results[pymongo_databases_key] = database_count
                    results[pymongo_collections_key] = collection_count

                if mma_file is not None:
                    lines = FS.read_lines(mma_file)
                    # lines = FS.read_encoded_lines(mma_file, encoding='utf-16le')  # from PowerShell cmd | Out-File -FilePath
                    for line in lines:
                        if 'Assessment Succeeded' in line:
                            results[mma_successful_key] = 'true'
                            print('Assessment Succeeded for {}'.format(cluster_name))
                        if 'Assessment Failed' in line:
                            results[mma_successful_key] = 'false'
                            print('Assessment Failed for {}'.format(cluster_name))

            FS.write_json(results, 'tmp/scan_mmaout_results.json')
            #scan_mmaout_results_report(clusters, results)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def scan_mmaout_results_report(cls):
        tr = TaskResult('scan_mmaout_results_report')
        try:
            Datasets.reset()
            clusters = FS.read_json('current/excel_clusters.json')
            results = FS.read_json('tmp/scan_mmaout_results.json')

            csv_lines, mma_error_files = list(), list()
            csv_lines.append(
                'cluster_name,mma_present,mma_successful,pymongo_present,pymongo_databases,pymongo_collections,note')
            for cluster_name in sorted(clusters.keys()):
                mma_present = cls.result_matching_value(results, cluster_name, 'mma_present')
                mma_successful = cls.result_matching_value(results, cluster_name, 'mma_successful').lower()
                pymongo_present = cls.result_matching_value(results, cluster_name, 'pymongo_present')
                pymongo_databases = cls.result_matching_value(results, cluster_name, 'pymongo_databases')
                pymongo_collections = cls.result_matching_value(results, cluster_name, 'pymongo_collections')
                note = ''
                if mma_successful == 'false':
                    if pymongo_present != 'None':
                        note = 'MMA error'
                        mma_error_files.append(mma_present)
                    else:
                        note = 'connection string issue'
                elif mma_successful == 'true':
                    if pymongo_present != 'None':
                        note = 'ok'
                        if int(pymongo_collections) < 1:
                            note = 'no collections'
                        if int(pymongo_databases) < 1:
                            note = 'no databases'
                csv_values = list()
                csv_values.append(cluster_name)
                csv_values.append(mma_present)
                csv_values.append(mma_successful)
                csv_values.append(pymongo_present)
                csv_values.append(pymongo_databases)
                csv_values.append(pymongo_collections)
                csv_values.append(note)
                csv_line = ','.join(csv_values)
                csv_lines.append(csv_line)
            FS.write_lines(csv_lines, 'tmp/scan_mmaout_results_report.csv')
            FS.write_lines(mma_error_files, 'tmp/mma_error_files.txt')

            mma_error_details = list()
            for mma_error_file in mma_error_files:
                mma_error_details.append(
                    '================================================================================')
                mma_error_details.append('file {}'.format(mma_error_file))
                mma_error_details.append('')
                lines = FS.read_lines(mma_error_file)
                for line in lines:
                    mma_error_details.append(line.rstrip())
                mma_error_details.append('')

            FS.write_lines(mma_error_details, 'tmp/mma_error_details.txt')
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def result_matching_value(results, cluster_name, name):
        for key in results.keys():
            if key.startswith(cluster_name):
                if name in key:
                    return str(results[key])
        return ''
    @classmethod
    def docscan_results_report(cls):
        tr = TaskResult('docscan_results_report')
        try:
            docscan_file_ojects = FS.walk('current/docscan')
            collections = list()
            clusters_largest = dict()  # key = name, value = DocscanCluster
            two_mb = Bytes.megabyte() * 2.0

            tr.log('docscan_file count: {}'.format(len(docscan_file_ojects)))
            for file_obj in docscan_file_ojects:
                abspath = file_obj['abspath']
                if abspath.endswith('.json'):
                    cluster_data = FS.read_json_utf8(abspath)
                    dcr = DocscanClusterResult(cluster_data)
                    dcr.parse()
                    cluster_name = dcr.cluster_name
                    dc = DocscanCluster(cluster_name)
                    clusters_largest[cluster_name] = dc
                    dcr.parse()
                    for dcoll in dcr.collections:
                        collections.append(dcoll)
                        dc.add_container(dcoll)
            tr.add_message('{} total collections'.format(len(collections)))

            collections.sort(key=lambda x: x.largest_doc_size)
            collections.reverse()

            collections_dict = dict()
            for c in collections:
                key = c.key
                collections_dict[key] = c.get_data()
            FS.write_json(collections_dict, 'current/docscan_merged_collections_dict.json')

            # create the docscan_by_largest_doc_in_collection report
            csv_lines = list()
            for idx, dc in enumerate(collections):
                if idx == 0:
                    csv_lines.append(dc.as_csv_header())
                csv_lines.append(dc.as_csv())
            FS.write_lines(csv_lines, 'current/docscan_by_largest_doc_in_collection.csv')

            # create the docscan_by_largest_doc_in_cluster report
            csv_lines = list()
            two_mb = Bytes.megabyte() * 2.0
            csv_lines.append('cluster_name,largest_doc_size,over_2mb')
            for cluster_name in sorted(clusters_largest.keys()):
                dc = clusters_largest[cluster_name]
                largest_size = dc.largest_doc_size
                over_2mb = 'no'
                if largest_size > two_mb:
                    over_2mb = 'yes'
                csv_lines.append('{},{},{}'.format(cluster_name,largest_size, over_2mb))
            FS.write_lines(csv_lines, 'current/docscan_by_largest_doc_in_cluster.csv')
            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr

    @classmethod
    def docscan_doc_capture(cls):
        tr = TaskResult('docscan_doc_capture')
        try:
            specs = FS.read_json('current/docscan_capture.json')
            for spec in specs:
                dbname   = spec['dbname']
                cname    = spec['cname']
                doc_id   = spec['doc_id']
                conn_str = spec['conn_str']
                capture  = spec['capture']
                tr.log('docscan_doc_capture; dbname: {}, cname: {}, doc_id: {}, conn_str: {}'.format(
                    dbname, cname, doc_id, conn_str))

                outfile = 'tmp/captured_docs/capture-doc-{}-{}-{}.json'.format(dbname, cname, doc_id)
                coordinates = '{}|{}|{}|{}'.format(dbname, cname, doc_id, conn_str)

                if capture.lower() == 'y':
                    opts = dict()
                    opts['conn_string'] = conn_str
                    opts['verbose'] = False
                    m = Mongo(opts)
                    m.set_db(dbname)
                    m.set_coll(cname)
                    try:
                        # first try a find_by_id with doc_id as an ObjectId
                        doc = m.find_by_id(doc_id)
                        jstr = json_util.dumps(doc)
                        obj = json.loads(jstr)
                        obj['__coordinates__'] = coordinates
                        obj['__json_size__'] = len(jstr)
                        FS.write_json(obj, outfile)
                        tr.log('captured {}'.format(outfile))
                    except:
                        # next try a find_by_id with doc_id as a string
                        query = dict()
                        query['_id'] = doc_id
                        doc = m.find_one(query)
                        jstr = json_util.dumps(doc)
                        obj = json.loads(jstr)
                        obj['__coordinates__'] = coordinates
                        obj['__json_size__'] = len(jstr)
                        FS.write_json(obj, outfile)
                        tr.log('captured {}'.format(outfile))

            tr.set_success(True)
        except Exception as e:
            tr.log(str(e))
            tr.log(traceback.format_exc())
            tr.set_success(False)
        return tr


    # private methods below

    @classmethod
    def _find_file_in_list(cls, filenames, cluster_name, suffix):
        for filename in filenames:
            if cluster_name in filename:
                if suffix in filename:
                    return filename
        return None

    @classmethod
    def __parse_assessment_name_and_id(cls, file_info):
        name, id = None, None
        abspath = file_info['abspath'].strip()

        if 'macos' in System.platform_info().lower():
            tokens = abspath.split('/')
        else:
            tokens = abspath.split("\\")

        for idx, token in enumerate(tokens):
            if token == 'MongoMigrationAssessment':
                id = tokens[idx + 1]

        lines = FS.read_lines(abspath)
        for line in lines:
            stripped = line.strip()
            if stripped.startswith('Assessment Name :'):
                tokens = stripped.split(':')
                name = tokens[1].strip()
        return (name, id)

    @classmethod
    def __lookup_cluster_name(cls, cluster_mapping, abspath):
        if 'macos' in System.platform_info().lower():
            tokens = abspath.split('/')
        else:
            tokens = abspath.split("\\")

        for idx, token in enumerate(tokens):
            if token == 'MongoMigrationAssessment':
                id = tokens[idx + 1]
                if id in cluster_mapping.keys():
                    return cluster_mapping[id]
        return '?'

    @classmethod
    def __verify_mongo_migration_assessment_report_html_file(cls, dir, mma_output_files):
        found = False
        for file_idx, file_info in enumerate(mma_output_files):
            abspath = file_info['abspath']
            if dir in abspath:
                if abspath.strip().endswith('mongo_migration_assessment_report.html'):
                    found = True
        return found


# Instances of class TaskResult are returned by the methods in class Tasks, above.
class TaskResult(object):

    def __init__(self, name):
        self.data = dict()
        self.data['name'] = name
        self.data['infiles'] = list()
        self.data['outfiles'] = list()
        self.data['messages'] = list()
        self.data['successful'] = False
        self.data['resp_obj'] = None

    def add_infile(self, filename):
        if filename is not None:
            self.data['infiles'].append(filename)

    def add_outfile(self, filename):
        if filename is not None:
            self.data['outfiles'].append(filename)

    def add_message(self, msg):
        if msg is not None:
            self.data['messages'].append(msg)

    def log(self, msg):
        print(msg)
        self.add_message(msg)

    def set_success(self, b):
        self.data['successful'] = b

    def get_resp_obj(self):
        return self.data['resp_obj']

    def set_resp_obj(self, obj):
        self.data['resp_obj'] = obj

    def display(self):
        print(json.dumps(self.data, sort_keys=False, indent=2))
